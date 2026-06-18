"""Build real, editable .xlsx workbooks from a structured spec (openpyxl).

Spec:
    {
      "sheets": [
        {"name": "Revenue",
         "headers": ["Quarter", "Revenue", "Churn"],
         "rows": [["Q1", 10, 0.04], ["Q2", 14, 0.03], ["Q3", 19, 0.02]],
         "chart": {"kind": "bar", "title": "Revenue",
                   "categories_col": 1, "values_col": 2}}   # 1-indexed columns
      ]
    }
"""

from __future__ import annotations

from typing import Any


def build_workbook(spec: dict[str, Any], out_path: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Font
    except Exception as exc:  # pragma: no cover - optional dep
        raise RuntimeError("openpyxl not installed (pip install openpyxl)") from exc

    sheets = spec.get("sheets", [])
    if not sheets:
        raise ValueError("workbook spec requires at least one sheet")

    wb = Workbook()
    wb.remove(wb.active)  # start clean

    for sheet in sheets:
        ws = wb.create_sheet(title=str(sheet.get("name", "Sheet"))[:31])
        headers = sheet.get("headers", [])
        if headers:
            ws.append(list(headers))
            for cell in ws[1]:
                cell.font = Font(bold=True)
        for row in sheet.get("rows", []):
            ws.append(list(row))

        chart_spec = sheet.get("chart")
        if chart_spec and headers and sheet.get("rows"):
            _add_chart(ws, chart_spec, BarChart, LineChart, Reference, len(sheet["rows"]))

    wb.save(out_path)
    return out_path


def _add_chart(ws, chart_spec, BarChart, LineChart, Reference, n_rows) -> None:
    chart = LineChart() if chart_spec.get("kind") == "line" else BarChart()
    chart.title = chart_spec.get("title", "")
    vcol = int(chart_spec.get("values_col", 2))
    ccol = int(chart_spec.get("categories_col", 1))
    data = Reference(ws, min_col=vcol, min_row=1, max_row=n_rows + 1)
    cats = Reference(ws, min_col=ccol, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F2")
